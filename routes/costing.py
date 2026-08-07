from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, literal_column
from sqlalchemy.dialects.mysql import insert as mysql_insert
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from database import get_db
from models import (
    PointsDb,
    ProjectCosts,
    ReconcileHistory,
    ProjectCostUpsert,
    ProjectCostRow,
    UploadResult,
    UploadRowError,
    TrendPoint,
)
from dependencies import get_costing_user

router = APIRouter(prefix="/api/costing", tags=["Costing"])

FULCRUM_ID = 23

PLACEHOLDER_SURVEY_NAME = "RDR00XXXX"
PLACEHOLDER_NOTES = "Please add any additional information"

TEMPLATE_HEADERS: list[str] = [
    "SurveyName",
    "Revenue (£)",
    "Translation Cost (£)",
    "Researcher Cost (£)",
    "Additional Costs (£)",
    "Notes",
]


def fulcrum_spend_expr():
    return case(
        (PointsDb.supplier == FULCRUM_ID, (PointsDb.cpi / 100.0 + 0.17) * 1.05),
        else_=PointsDb.cpi / 100.0,
    )


def _sample_spend_map(db: Session, survey_names: list[str] | None = None) -> dict[str, float]:
    q = db.query(
        PointsDb.project.label("survey_name"),
        func.sum(fulcrum_spend_expr()).label("spend"),
    ).filter(PointsDb.status == 1)
    if survey_names is not None:
        if not survey_names:
            return {}
        q = q.filter(PointsDb.project.in_(survey_names))
    rows = q.group_by(PointsDb.project).all()
    return {r.survey_name: float(r.spend or 0) for r in rows}


def _reconciled_set(db: Session, survey_names: list[str] | None = None) -> set[str]:
    q = db.query(ReconcileHistory.surveyid).distinct()
    if survey_names is not None:
        if not survey_names:
            return set()
        q = q.filter(ReconcileHistory.surveyid.in_(survey_names))
    return {r.surveyid for r in q.all()}

def _pl1_survey_set(db: Session, survey_names: list[str] | None = None) -> set[str]:
    q = db.query(PointsDb.project).filter(
        PointsDb.supplier == FULCRUM_ID, PointsDb.status == 1
    ).distinct()
    if survey_names is not None:
        if not survey_names:
            return set()
        q = q.filter(PointsDb.project.in_(survey_names))
    return {r.project for r in q.all()}


def _to_row(pc: ProjectCosts, sample: float, reconciled: bool) -> ProjectCostRow:
    revenue = float(pc.revenue_gbp)
    translations = float(pc.translations_gbp)
    researcher = float(pc.researcher_cost_gbp)
    additional = float(pc.additional_outgoing_gbp)
    outgoings = sample + translations + researcher + additional
    net = revenue - outgoings
    margin = (net / revenue * 100) if revenue else None
    return ProjectCostRow(
        survey_name=pc.survey_name,
        revenue_gbp=revenue,
        sample_cost_gbp=sample,
        translations_gbp=translations,
        researcher_cost_gbp=researcher,
        additional_outgoing_gbp=additional,
        total_outgoings_gbp=outgoings,
        net_gbp=net,
        margin_pct=margin,
        notes=pc.notes,
        uploaded_by=pc.uploaded_by,
        uploaded_at=pc.uploaded_at.isoformat() if pc.uploaded_at else None,
        is_reconciled=reconciled,
    )


@router.get("", response_model=list[ProjectCostRow])
def list_project_costs(
    db: Session = Depends(get_db),
    user: dict = Depends(get_costing_user),
):
    rows = db.query(ProjectCosts).order_by(ProjectCosts.uploaded_at.desc()).all()
    names = [r.survey_name for r in rows]
    sample_map = _sample_spend_map(db, names)
    reconciled = _reconciled_set(db, names)
    pl1_surveys = _pl1_survey_set(db, names)
    return [
        _to_row(r, sample_map.get(r.survey_name, 0.0), (r.survey_name in reconciled) or (r.survey_name not in pl1_surveys))
        for r in rows
    ]


@router.get("/{survey_name}", response_model=ProjectCostRow)
def get_project_cost(
    survey_name: str,
    db: Session = Depends(get_db),
    user: dict = Depends(get_costing_user),
):
    pc = db.query(ProjectCosts).filter(ProjectCosts.survey_name == survey_name).first()
    if not pc:
        raise HTTPException(status_code=404, detail="Not found")
    sample = _sample_spend_map(db, [survey_name]).get(survey_name, 0.0)
    reconciled = (survey_name in _reconciled_set(db, [survey_name])) or (survey_name not in _pl1_survey_set(db, [survey_name]))
    return _to_row(pc, sample, reconciled)


@router.post("", response_model=ProjectCostRow)
def upsert_project_cost(
    payload: ProjectCostUpsert,
    db: Session = Depends(get_db),
    user: dict = Depends(get_costing_user),
):
    name = payload.survey_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="survey_name required")
    valid_names = {p[0] for p in db.query(PointsDb.project).distinct().all() if p[0]}
    valid_norm = {n.strip().lower(): n for n in valid_names}
    matched = valid_norm.get(name.lower())
    if not matched:
        raise HTTPException(
            status_code=400,
            detail="Survey not found in CPI. Costs can only be entered for surveys that have already been launched.",
        )
    name = matched
    stmt = mysql_insert(ProjectCosts).values(
        survey_name=name,
        revenue_gbp=payload.revenue_gbp,
        translations_gbp=payload.translations_gbp,
        researcher_cost_gbp=payload.researcher_cost_gbp,
        additional_outgoing_gbp=payload.additional_outgoing_gbp,
        notes=payload.notes,
        uploaded_by=user.get("email"),
        uploaded_at=datetime.utcnow(),
    )
    stmt = stmt.on_duplicate_key_update(
        revenue_gbp=stmt.inserted.revenue_gbp,
        translations_gbp=stmt.inserted.translations_gbp,
        researcher_cost_gbp=stmt.inserted.researcher_cost_gbp,
        additional_outgoing_gbp=stmt.inserted.additional_outgoing_gbp,
        notes=stmt.inserted.notes,
        uploaded_by=stmt.inserted.uploaded_by,
        uploaded_at=stmt.inserted.uploaded_at,
    )
    db.execute(stmt)
    db.commit()
    pc = db.query(ProjectCosts).filter(ProjectCosts.survey_name == name).first()
    sample = _sample_spend_map(db, [name]).get(name, 0.0)
    reconciled = (name in _reconciled_set(db, [name])) or (name not in _pl1_survey_set(db, [name]))
    return _to_row(pc, sample, reconciled)



@router.get("/template/download")
def download_template(user: dict = Depends(get_costing_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Cost"

    green_fill  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD34E", end_color="FFD34E", fill_type="solid")
    white_bold  = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center", vertical="center")
    middle      = Alignment(horizontal="left",   vertical="center")

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    
    ncols = len(TEMPLATE_HEADERS)
    last_col = chr(ord("A") + ncols - 1)  

    # Row 2 - SPPend Samples
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = "3GEM Project Cost Template"
    t.font      = Font(bold=True, size=14, color="000000")
    t.fill      = yellow_fill
    t.alignment = center
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).border = border
    ws.row_dimensions[1].height = 32

    # Row 2 — column headers (green, white bold)
    for col, heading in enumerate(TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=heading)
        c.fill      = green_fill
        c.font      = white_bold
        c.alignment = center
        c.border    = border
    ws.row_dimensions[2].height = 22

    # Row 3 — example placeholder row (users overwrite or delete)
    placeholder_row = [PLACEHOLDER_SURVEY_NAME, None, None, None, None, PLACEHOLDER_NOTES]
    for col, val in enumerate(placeholder_row, start=1):
        c = ws.cell(row=3, column=col, value=val)
        c.font      = Font(color="000000")
        c.alignment = middle
        c.border    = border
    ws.row_dimensions[3].height = 20

    # Column widths
    widths = [30, 16, 20, 20, 20, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
        
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{PLACEHOLDER_SURVEY_NAME}.xlsx"'},
    )


@router.post("/upload", response_model=UploadResult)
def upload_costs(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(get_costing_user),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Expected .xlsx file")
    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read workbook")
    ws = wb.active

    # Validate header row (row 2 must match TEMPLATE_HEADERS)
    header_row = [
        str(c.value or "").strip().lower()
        for c in next(ws.iter_rows(min_row=2, max_row=2, max_col=len(TEMPLATE_HEADERS)))
    ]
    expected = [h.lower() for h in TEMPLATE_HEADERS]
    if header_row != expected:
        raise HTTPException(
            status_code=400,
            detail="Header row does not match template. Re-download the template — row 2 must contain the standard column headers.",
        )

    # Preload valid survey names from points table
    valid_names = {p[0] for p in db.query(PointsDb.project).distinct().all() if p[0]}
    valid_norm = {n.strip().lower(): n for n in valid_names}

    errors: list[UploadRowError] = []
    rows_to_save: list[tuple[str, float, float, float, float, str | None]] = []
    seen_in_file: set[str] = set()

    def to_float(val, row_idx: int, col: str, name: str | None, required: bool = False) -> float | None:
        if val is None or str(val).strip() == "":
            if required:
                errors.append(UploadRowError(row=row_idx, survey_name=name, error=f"{col} is required"))
                return None
            return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            errors.append(UploadRowError(row=row_idx, survey_name=name, error=f"{col} must be a number"))
            return None

    # Iterate data rows from row 3
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=3, max_col=len(TEMPLATE_HEADERS), values_only=True),
        start=3,
    ):
        name_raw, rev_raw, trans_raw, res_raw, add_raw, notes_raw = row

        # Skip fully-blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        # Skip the placeholder example row
        if name_raw and str(name_raw).strip().lower() == PLACEHOLDER_SURVEY_NAME.lower():
            continue

        if not name_raw or str(name_raw).strip() == "":
            errors.append(UploadRowError(row=row_idx, survey_name=None, error="SurveyName is empty"))
            continue

        matched = valid_norm.get(str(name_raw).strip().lower())
        if not matched:
            errors.append(UploadRowError(
                row=row_idx, survey_name=str(name_raw),
                error="Survey name not found in points",
            ))
            continue

        if matched in seen_in_file:
            errors.append(UploadRowError(
                row=row_idx, survey_name=matched,
                error="Duplicate survey within this file",
            ))
            continue
        seen_in_file.add(matched)

        revenue      = to_float(rev_raw,   row_idx, "Revenue (£)",          matched, required=True)
        translations = to_float(trans_raw, row_idx, "Translation Cost (£)", matched)
        researcher   = to_float(res_raw,   row_idx, "Researcher Cost (£)",  matched)
        additional   = to_float(add_raw,   row_idx, "Additional Costs (£)", matched)

        if revenue is None or translations is None or researcher is None or additional is None:
            continue  # already logged in to_float

        if revenue <= 0:
            errors.append(UploadRowError(row=row_idx, survey_name=matched, error="Revenue must be > 0"))
            continue

        notes: str | None = None
        if notes_raw is not None:
            stripped = str(notes_raw).strip()
            if stripped and stripped.lower() != PLACEHOLDER_NOTES.lower():
                notes = stripped

        rows_to_save.append((matched, revenue, translations, researcher, additional, notes))

    # Reject the whole file if any row had an error
    if errors:
        return UploadResult(inserted=0, updated=0, errors=errors)

    if not rows_to_save:
        return UploadResult(
            inserted=0, updated=0,
            errors=[UploadRowError(
                row=3, survey_name=None,
                error="No data rows found (file appears empty or only contains the placeholder).",
            )],
        )

    # Figure out which names already exist (for accurate inserted/updated counts)
    existing_names = {
        r[0] for r in db.query(ProjectCosts.survey_name)
        .filter(ProjectCosts.survey_name.in_([rs[0] for rs in rows_to_save]))
        .all()
    }

    now = datetime.utcnow()
    inserted = 0
    updated = 0
    for name, rev, trans, res, add, notes in rows_to_save:
        is_update = name in existing_names
        stmt = mysql_insert(ProjectCosts).values(
            survey_name=name,
            revenue_gbp=rev,
            translations_gbp=trans,
            researcher_cost_gbp=res,
            additional_outgoing_gbp=add,
            notes=notes,
            uploaded_by=user.get("email"),
            uploaded_at=now,
        ).on_duplicate_key_update(
            revenue_gbp=rev,
            translations_gbp=trans,
            researcher_cost_gbp=res,
            additional_outgoing_gbp=add,
            notes=notes,
            uploaded_by=user.get("email"),
            uploaded_at=now,
        )
        db.execute(stmt)
        if is_update:
            updated += 1
        else:
            inserted += 1
    db.commit()

    return UploadResult(inserted=inserted, updated=updated, errors=[])


@router.get("/analytics/trend", response_model=list[TrendPoint])
def get_trend(
    db: Session = Depends(get_db),
    user: dict = Depends(get_costing_user),
):
    # Pull all project_costs rows
    cost_rows = db.query(ProjectCosts).all()
    if not cost_rows:
        return []

    names = [r.survey_name for r in cost_rows]

    # Sample spend per survey
    sample_map = _sample_spend_map(db, names)

    # Earliest complete date per survey (Excel-corrected: stime - 2 days)
    date_rows = (
        db.query(
            PointsDb.project,
            func.min(
                func.date_sub(PointsDb.stime, literal_column("INTERVAL 2 DAY"))
            ).label("first_date"),
        )
        .filter(PointsDb.project.in_(names), PointsDb.status == 1)
        .group_by(PointsDb.project)
        .all()
    )
    first_date_map: dict[str, str] = {}
    for dr in date_rows:
        if dr.first_date:
            d = dr.first_date
            first_date_map[dr.project] = f"{d.year}-{str(d.month).zfill(2)}"

    # Total completes per survey (for proportional allocation)
    completes_rows = (
        db.query(
            PointsDb.project,
            func.count().label("total"),
        )
        .filter(PointsDb.project.in_(names), PointsDb.status == 1)
        .group_by(PointsDb.project)
        .all()
    )
    completes_map = {r.project: r.total for r in completes_rows}

    # Build monthly buckets over last 12 months
    from datetime import date, timedelta
    today = date.today()
    months: list[str] = []
    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{str(m).zfill(2)}")

    revenue_buckets: dict[str, float] = {mo: 0.0 for mo in months}
    outgoings_buckets: dict[str, float] = {mo: 0.0 for mo in months}

    for pc in cost_rows:
        mo = first_date_map.get(pc.survey_name)
        if not mo or mo not in revenue_buckets:
            continue
        revenue = float(pc.revenue_gbp)
        sample = sample_map.get(pc.survey_name, 0.0)
        other = float(pc.translations_gbp) + float(pc.researcher_cost_gbp) + float(pc.additional_outgoing_gbp)
        outgoings = sample + other
        revenue_buckets[mo] += revenue
        outgoings_buckets[mo] += outgoings

    result: list[TrendPoint] = []
    for mo in months:
        y, m = mo.split("-")
        label_date = date(int(y), int(m), 1)
        label = label_date.strftime("%b %y")
        result.append(TrendPoint(
    month=mo,
    label=label,
    revenue=revenue_buckets[mo],
    outgoings=outgoings_buckets[mo],
    net=revenue_buckets[mo] - outgoings_buckets[mo],
))

    return result